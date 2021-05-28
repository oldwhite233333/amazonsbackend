import openpyxl
from fpdf import FPDF
from PIL import Image


class Dao:
    def __init__(self, eanlist, amount):
        self.eanlist = eanlist
        self.amount = amount
        book = openpyxl.load_workbook('products.xlsx', data_only=True)
        dict = {}
        sheet = book.worksheets[0]
        rows = iter(sheet.rows)
        next(rows)
        self.productList = []
        for row in rows:
            product = Product(row)
            self.productList.append(product)
            dict[row[3].value] = product
        products = []
        boxs = []
        ean_to_asin = {}
        for i in self.productList:
            ean_to_asin[i.ean] = i.asin
        for i in eanlist:
            products.append(dict[i.strip()])
        for z in amount:
            boxs.append(int(z))
        for i in range(0, len(boxs)):
            boxs[i] = round(boxs[i] / products[i].unit_per_box)
        self.products = products
        self.boxs = boxs
        self.ean_to_asin = ean_to_asin


def to_inch(arg):
    return round(arg * 0.39, 2)


class Spec:
    def __init__(self, length, width, height, weight):
        self.length = length
        self.width = width
        self.height = height
        self.weight = weight

    def __str__(self):
        return str(self.get_volume()) + " " + str(self.weight)

    def get_volume(self):
        return round(self.width * self.height * self.length / (100 * 100 * 100), 3)

    def get_inch(self):
        return to_inch(self.length), to_inch(self.width), to_inch(self.height)

    def get_pound(self):
        return round(self.weight * 2.2, 2)


class Product:
    def __init__(self, row, no="xlsx"):
        if no == "":
            self.no = no
            self.name = row.name
            self.asin = row.asin
            self.model_id = row.model_id
            self.ean = row.ean
            self.rmb_price = row.rmb_price
            self.usd_price = row.rmb_price / 6.5 if self.rmb_price else 0
            self.ship_fee = row.ship_fee
            self.item_amount = row.item_amount
            self.inner_spec = Spec(row.inner_length, row.inner_width, row.inner_height, row.inner_weight)
            self.out_spec = Spec(row.out_length, row.out_width, row.out_height, row.out_weight)
            self.unit_per_box = row.unit_per_box
        else:
            self.row = row
            self.name = self.v('a')
            self.asin = self.v('b')
            self.model_id = self.v('c')
            self.ean = self.v('d')
            self.rmb_price = self.v('e')
            self.usd_price = self.rmb_price / 6.5 if self.rmb_price else 0
            self.ship_fee = self.v('g')
            self.item_amount = self.v('h')
            self.inner_spec = Spec(self.v('i'), self.v('j'), self.v('k'), self.v('l'))
            self.out_spec = Spec(self.v('p'), self.v('q'), self.v('r'), self.v('s'))
            self.unit_per_box = self.v('x')

    def v(self, s):
        return self.row[ord(s) - 97].value


def write_products(rowNum, product):
    return


def write_to_doc(products, boxs):
    book = openpyxl.load_workbook('result.xlsx')
    sheet = book.worksheets[0]
    for i in range(0, len(boxs)):
        sheet.cell(row=i + 1, column=1, value=products[i].unit_per_box)
        sheet.cell(row=i + 1, column=2, value=boxs[i])
        sheet.cell(row=i + 1, column=3, value=products[i].out_spec.get_volume())
        sheet.cell(row=i + 1, column=4, value=products[i].out_spec.weight)
        sheet.cell(row=i + 1, column=5, value=products[i].out_spec.get_volume() * boxs[i])
        sheet.cell(row=i + 1, column=6, value=products[i].out_spec.weight * boxs[i])
    book.save('result.xlsx')


def write_product_xlsx(product: Product, row):
    # row+=1
    row = 104
    book = openpyxl.load_workbook('products.xlsx')
    sheet = book.worksheets[0]

    def s(i):
        return ord(i) - 96

    sheet.cell(row, column=s('a'), value=product.name)
    sheet.cell(row, column=s('b'), value=product.asin)
    sheet.cell(row, column=s('c'), value=product.model_id)
    sheet.cell(row, column=s('d'), value=str(product.ean))
    sheet.cell(row, column=s('e'), value=product.rmb_price)
    sheet.cell(row, column=s('f'), value=round(product.rmb_price / 6.5, 2))
    sheet.cell(row, column=s('g'), value=product.ship_fee)
    sheet.cell(row, column=s('h'), value=product.item_amount)
    sheet.cell(row, column=s('i'), value=product.inner_spec.length)
    sheet.cell(row, column=s('j'), value=product.inner_spec.width)
    sheet.cell(row, column=s('k'), value=product.inner_spec.height)
    sheet.cell(row, column=s('l'), value=product.inner_spec.weight)
    sheet.cell(row, column=s('m'), value=product.inner_spec.get_inch()[0])
    sheet.cell(row, column=s('n'), value=product.inner_spec.get_inch()[1])
    sheet.cell(row, column=s('o'), value=product.inner_spec.get_inch()[2])
    sheet.cell(row, column=s('p'), value=product.out_spec.length)
    sheet.cell(row, column=s('q'), value=product.out_spec.width)
    sheet.cell(row, column=s('r'), value=product.out_spec.height)
    sheet.cell(row, column=s('s'), value=product.out_spec.weight)
    sheet.cell(row, column=s('t'), value=product.out_spec.get_inch()[0])
    sheet.cell(row, column=s('u'), value=product.out_spec.get_inch()[1])
    sheet.cell(row, column=s('v'), value=product.out_spec.get_inch()[2])
    sheet.cell(row, column=s('w'), value=product.out_spec.get_pound())
    sheet.cell(row, column=s('x'), value=product.unit_per_box)
    sheet.cell(row, column=s('y'), value=product.out_spec.get_volume())
    book.save('products.xlsx')


def write_doc_xiongda(products, boxs):
    book = openpyxl.load_workbook('result.xlsx')
    sheet = book.worksheets[0]
    for i in range(0, len(boxs)):
        sheet.cell(row=i + 1, column=1, value=boxs[i])
        sheet.cell(row=i + 1, column=2, value=boxs[i] * products[i].unit_per_box)
        sheet.cell(row=i + 1, column=6, value=products[i].out_spec.weight * boxs[i])
        sheet.cell(row=i + 1, column=7, value=products[i].out_spec.length)
        sheet.cell(row=i + 1, column=8, value=products[i].out_spec.width)
        sheet.cell(row=i + 1, column=9, value=products[i].out_spec.height)
    book.save('result.xlsx')
    return 0;


def write_doc_plan(products, boxs, ean_to_asin, addr, planName):
    plan_padding = 12
    txtdir = 'inbound.txt'
    if addr != 'hangzhou':
        txtdir = 'inbound_xiongda.txt'
    plan = [i.split('\t') for i in open(txtdir)]
    plan[0][1] = planName
    for i in range(0, len(boxs)):
        plan[plan_padding + i][0] = ean_to_asin.get(products[i].ean, products[i].ean)
        plan[plan_padding + i][1] = str(products[i].unit_per_box)
        plan[plan_padding + i][2] = str(boxs[i])
        plan[plan_padding + i][3] = str(boxs[i] * products[i].unit_per_box)
    f = open('./shipment.txt', 'wb+')
    for i in plan:
        bet = '\t'
        f.write(bet.join(i).encode())
    f.close()


def write_doc_shippment(dao, addr):
    products = dao.products
    boxs = dao.boxs
    book = openpyxl.load_workbook(addr)
    sheet = book.worksheets[0]
    # sheet.cell(row=5,column=14,value=9527);
    row_padding = 5
    col_padding = 14
    box_padding = 0
    spec_paading = row_padding + len(products) + 2
    for i in range(len(products)):
        for z in range(0, boxs[i]):
            sheet.cell(row=row_padding + i, column=col_padding + box_padding, value=products[i].unit_per_box)
            sheet.cell(row=spec_paading, column=col_padding + box_padding, value=products[i].out_spec.get_pound())
            inch = products[i].out_spec.get_inch()
            sheet.cell(row=spec_paading + 1, column=col_padding + box_padding, value=inch[0])
            sheet.cell(row=spec_paading + 2, column=col_padding + box_padding, value=inch[1])
            sheet.cell(row=spec_paading + 3, column=col_padding + box_padding, value=inch[2])
            box_padding += 1
            # sheet.cell(row=14,column=col_padding+z,value=products[i].out_spec.get_volume())
            # sheet.cell()
    # for i in range(len(products)):
    # print(products[i].unit_per_box, boxs[i])
    # print(products[i].out_spec.get_pound())
    # print(products[i].out_spec.get_inch())
    # print('==================')
    book.save('shipment.xlsx')
    return


def write_doc_shenzhen(products, boxs):
    # print(sum(boxs))
    book = openpyxl.load_workbook('result.xlsx')
    sheet = book.worksheets[0]
    count = 1;
    # print(boxs)
    for i in range(len(boxs)):
        for z in range(boxs[i]):
            # #print(i)
            p = products[i]
            sheet.cell(row=count, column=1, value=p.unit_per_box)
            sheet.cell(row=count, column=2, value=p.out_spec.weight)
            sheet.cell(row=count, column=3, value=p.out_spec.length)
            sheet.cell(row=count, column=4, value=p.out_spec.width)
            sheet.cell(row=count, column=5, value=p.out_spec.height)
            sheet.cell(row=count, column=6, value=p.inner_spec.weight / 1000)
            count += 1
    book.save('result.xlsx')
    return


def getInnerSpec(eanList):
    # print(eanList)
    for i in eanList:
        i = i.strip('\n')
        # print(i)
        # print(dict[i].inner_spec.length)
        # print(dict[i].inner_spec.width)
        # print(dict[i].inner_spec.height)
        # print(dict[i].inner_spec.weight)
        # print("==============")


def getEanBySku(eanlist, ean_to_asin):
    # print(eanlist)
    for z in eanlist:
        exist = False
        for i in ean_to_asin.keys():
            if ean_to_asin[i] == z:
                exist = True
                # print(i)
        # if (not exist): #print(z)
    # for i in eanlist:
    #     #print(list(ean_to_asin.keys())[list(ean_to_asin.values()).ean_to_asin(i)])


def dao_init(eanlist, amount):
    book = openpyxl.load_workbook('test.xlsx')
    dict = {}
    sheet = book.worksheets[0]
    for row in sheet.rows:
        innerSpec = Spec(row[6].value, row[7].value, row[8].value, row[9].value)
        outerSpec = Spec(row[13].value, row[14].value, row[15].value, row[16].value)
        product = Product(row[1].value,
                          innerSpec, outerSpec, row[21].value)

        dict[row[1].value] = product
    products = []
    boxs = []
    for i in open('ean2sku.csv'):
        k = i.split(',')
        if (i == 'SKU'): continue
        # ean_to_asin[k[1]] = k[0]
    for i in eanlist:
        if i != '':
            products.append(dict[i.strip()])
    for z in amount:
        boxs.append(int(z))
    for i in range(0, len(boxs)):
        boxs[i] = round(boxs[i] / products[i].unit_per_box)
    write_doc_plan(products, boxs)


def getFbaEanAndAmount(addr):
    book = openpyxl.load_workbook(addr)
    sheet = book.worksheets[0]
    ean_padding_col = 5
    ean_padding_row = 4
    cur = 0
    eanlist = []
    qualityList = []
    while 1:
        ean = sheet[ean_padding_col + cur][ean_padding_row].value
        if ean == None: break
        ean = ean.replace('EAN: ', '')
        quality = sheet[ean_padding_col + cur][ean_padding_row + 4].value
        cur += 1
        eanlist.append(ean)
        qualityList.append(quality)
    return (eanlist, qualityList)


def makePdf(ean):
    ean = [i.replace('\n', '') + '-new.png' for i in ean]
    pdfFileName = 'label.pdf'
    cover = Image.open('./madeinchina/' + ean[0])
    width, height = cover.size
    pdf = FPDF(unit="pt", format=[width, height])
    for page in ean:
        try:
            pdf.add_page()
            pdf.image('./madeinchina/' + page, 0, 0)
        except Exception as e:
            continue
    pdf.output(pdfFileName, "F")


def makeWorkFlowPlan(dao:Dao):
    def s(i):
        return ord(i) - 96
    row_padding=9
    book = openpyxl.load_workbook("workflow_template.xlsx")
    sheet = book.worksheets[2]
    for i in range(len(dao.products)):
        print(i)
        p=dao.products[i]
        out_spec=p.out_spec.get_inch()
        pound=p.out_spec.get_pound()
        sheet.cell(row_padding+i,s('a'),value=p.asin)
        sheet.cell(row_padding+i,s('b'),value=dao.amount[i])
        sheet.cell(row_padding+i,s('f'),value=p.unit_per_box)
        sheet.cell(row_padding+i,s('g'),value=dao.boxs[i])
        sheet.cell(row_padding+i,s('h'),value=out_spec[0])
        sheet.cell(row_padding+i,s('i'),value=out_spec[1])
        sheet.cell(row_padding+i,s('j'),value=out_spec[2])
        sheet.cell(row_padding+i,s('k'),value=pound)
    book.save("workflow.xlsx")

# if __name__ == '__main__':
#     makeWorkFlowPlan()

